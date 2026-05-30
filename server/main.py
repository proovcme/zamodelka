import uvicorn
import io
import math
from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, ConfigDict
from typing import Optional, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .database import engine, Base, get_db
from .models import Project, Sortament

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI(title="Vent MVP Backend API", version="1.0.0")

# Setup CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic schemas
class ProjectCreate(BaseModel):
    name: str
    units: Optional[str] = "mm"

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    version: Optional[int] = None
    units: Optional[str] = None
    graph: Optional[Dict[str, Any]] = None

class ProjectResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: str
    name: str
    version: int
    units: str
    graph: Dict[str, Any]

class SortamentResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    ref: str
    shape: str
    d: Optional[float] = None
    w: Optional[float] = None
    h: Optional[float] = None
    wall_thickness: Optional[float] = None
    mass_per_m: Optional[float] = None
    source: Optional[str] = None

# Seed Sortament data on startup
@app.on_event("startup")
def seed_sortament():
    db = next(get_db())
    try:
        count = db.query(Sortament).count()
        if count == 0:
            initial_sortament = [
                # Round ducts (ВСН 353-86)
                Sortament(ref="VSN353-R-100", shape="round", d=100.0, wall_thickness=0.5, mass_per_m=1.2, source="ВСН 353-86"),
                Sortament(ref="VSN353-R-125", shape="round", d=125.0, wall_thickness=0.5, mass_per_m=1.5, source="ВСН 353-86"),
                Sortament(ref="VSN353-R-160", shape="round", d=160.0, wall_thickness=0.5, mass_per_m=1.9, source="ВСН 353-86"),
                Sortament(ref="VSN353-R-200", shape="round", d=200.0, wall_thickness=0.5, mass_per_m=2.4, source="ВСН 353-86"),
                Sortament(ref="VSN353-R-250", shape="round", d=250.0, wall_thickness=0.5, mass_per_m=3.1, source="ВСН 353-86"),
                Sortament(ref="VSN353-R-315", shape="round", d=315.0, wall_thickness=0.5, mass_per_m=3.9, source="ВСН 353-86"),
                Sortament(ref="VSN353-R-400", shape="round", d=400.0, wall_thickness=0.5, mass_per_m=4.9, source="ВСН 353-86"),
                
                # Rectangular ducts (ВСН 353-86)
                Sortament(ref="VSN353-REC-150x100", shape="rectangular", w=150.0, h=100.0, wall_thickness=0.5, mass_per_m=2.0, source="ВСН 353-86"),
                Sortament(ref="VSN353-REC-200x150", shape="rectangular", w=200.0, h=150.0, wall_thickness=0.5, mass_per_m=2.8, source="ВСН 353-86"),
                Sortament(ref="VSN353-REC-250x200", shape="rectangular", w=250.0, h=200.0, wall_thickness=0.5, mass_per_m=3.5, source="ВСН 353-86"),
                Sortament(ref="VSN353-REC-400x250", shape="rectangular", w=400.0, h=250.0, wall_thickness=0.6, mass_per_m=6.1, source="ВСН 353-86"),
            ]
            db.bulk_save_objects(initial_sortament)
            db.commit()
            print("Sortament seeded successfully.")

        # Досидка кабельных лотков
        tray_count = db.query(Sortament).filter(Sortament.shape == "tray").count()
        if tray_count == 0:
            tray_sortament = [
                Sortament(ref="TRAY-100x50", shape="tray", w=100.0, h=50.0, wall_thickness=1.0, mass_per_m=1.1, source="Сортамент лотков"),
                Sortament(ref="TRAY-200x80", shape="tray", w=200.0, h=80.0, wall_thickness=1.2, mass_per_m=1.8, source="Сортамент лотков"),
                Sortament(ref="TRAY-300x100", shape="tray", w=300.0, h=100.0, wall_thickness=1.5, mass_per_m=2.7, source="Сортамент лотков"),
                Sortament(ref="TRAY-400x100", shape="tray", w=400.0, h=100.0, wall_thickness=1.5, mass_per_m=3.5, source="Сортамент лотков"),
            ]
            db.bulk_save_objects(tray_sortament)
            db.commit()
            print("Trays sortament seeded.")

        # Досидка трубопроводов
        pipe_count = db.query(Sortament).filter(Sortament.shape == "pipe").count()
        if pipe_count == 0:
            pipe_sortament = [
                Sortament(ref="PIPE-15", shape="pipe", d=15.0, wall_thickness=2.5, mass_per_m=1.2, source="ГОСТ 3262-75"),
                Sortament(ref="PIPE-20", shape="pipe", d=20.0, wall_thickness=2.8, mass_per_m=1.7, source="ГОСТ 3262-75"),
                Sortament(ref="PIPE-25", shape="pipe", d=25.0, wall_thickness=3.2, mass_per_m=2.4, source="ГОСТ 3262-75"),
                Sortament(ref="PIPE-32", shape="pipe", d=32.0, wall_thickness=3.2, mass_per_m=3.1, source="ГОСТ 3262-75"),
                Sortament(ref="PIPE-40", shape="pipe", d=40.0, wall_thickness=3.5, mass_per_m=3.8, source="ГОСТ 3262-75"),
                Sortament(ref="PIPE-50", shape="pipe", d=50.0, wall_thickness=3.5, mass_per_m=4.9, source="ГОСТ 3262-75"),
                Sortament(ref="PIPE-80", shape="pipe", d=80.0, wall_thickness=4.0, mass_per_m=8.4, source="ГОСТ 3262-75"),
                Sortament(ref="PIPE-100", shape="pipe", d=100.0, wall_thickness=4.5, mass_per_m=11.6, source="ГОСТ 3262-75"),
            ]
            db.bulk_save_objects(pipe_sortament)
            db.commit()
            print("Pipes sortament seeded.")
    except Exception as e:
        db.rollback()
        print(f"Error seeding sortament: {e}")
    finally:
        db.close()

# API Endpoints
@app.get("/sortament", response_model=List[SortamentResponse])
def read_sortament(shape: Optional[str] = Query(None, description="Filter by shape: 'round' or 'rectangular'"), db: Session = Depends(get_db)):
    query = db.query(Sortament)
    if shape:
        query = query.filter(Sortament.shape == shape)
    return query.all()

@app.post("/projects", response_model=ProjectResponse)
def create_project(project_in: ProjectCreate, db: Session = Depends(get_db)):
    db_project = Project(
        name=project_in.name,
        units=project_in.units,
        graph={"elements": []}
    )
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    return db_project

@app.get("/projects/{project_id}", response_model=ProjectResponse)
def read_project(project_id: str, db: Session = Depends(get_db)):
    db_project = db.query(Project).filter(Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    return db_project

@app.put("/projects/{project_id}", response_model=ProjectResponse)
def update_project(project_id: str, project_in: ProjectUpdate, db: Session = Depends(get_db)):
    db_project = db.query(Project).filter(Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    update_data = project_in.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_project, key, value)
    
    db.commit()
    db.refresh(db_project)
    return db_project

# Масса погонного метра по сортаменту ВСН 353-86
DUCT_MASS_MAP = {
    "VSN353-R-100": 1.2,
    "VSN353-R-125": 1.5,
    "VSN353-R-160": 1.9,
    "VSN353-R-200": 2.4,
    "VSN353-R-250": 3.1,
    "VSN353-R-315": 3.9,
    "VSN353-R-400": 4.9,
    "VSN353-REC-150x100": 2.0,
    "VSN353-REC-200x150": 2.8,
    "VSN353-REC-250x200": 3.5,
    "VSN353-REC-400x250": 6.1,
}

@app.get("/projects/{project_id}/export/xlsx")
def export_project_xlsx(project_id: str, db: Session = Depends(get_db)):
    db_project = db.query(Project).filter(Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
        
    elements = db_project.graph.get("elements", [])
    
    # 1. Расчет ведомости (BOM)
    ducts_data = {}
    fittings_data = {}
    terminals_data = {}
    equipment_data = {}
    
    for elem in elements:
        el_type = elem.get("type")
        if el_type == "duct":
            start = elem.get("start", [0, 0, 0])
            end = elem.get("end", [0, 0, 0])
            dx = end[0] - start[0]
            dy = end[1] - start[1]
            dz = end[2] - start[2]
            length_m = math.sqrt(dx*dx + dy*dy + dz*dz) / 1000.0
            
            shape = elem.get("shape", "round")
            size = elem.get("size", {})
            
            if shape == "round":
                d = size.get("d", 200)
                size_label = f"⌀{int(d)}"
                area = math.pi * (d / 1000.0) * length_m
            else:
                w = size.get("w", 300)
                h = size.get("h", 200)
                size_label = f"{int(w)}x{int(h)}"
                area = 2 * ((w + h) / 1000.0) * length_m
                
            ref = elem.get("sortamentRef", "unknown")
            mass_per_m = DUCT_MASS_MAP.get(ref, (d / 100.0) * 1.2 if shape == "round" else ((w + h) / 200.0) * 2.0)
            weight = mass_per_m * length_m
            
            key = (shape, size_label, ref)
            if key in ducts_data:
                ducts_data[key]["length"] += length_m
                ducts_data[key]["area"] += area
                ducts_data[key]["weight"] += weight
            else:
                ducts_data[key] = {
                    "name": "Воздуховод круглый" if shape == "round" else "Воздуховод прямоугольный",
                    "size": size_label,
                    "length": length_m,
                    "area": area,
                    "weight": weight
                }
        elif el_type == "fitting":
            kind = elem.get("kind", "bend")
            size = elem.get("size", {})
            if "d" in size:
                size_label = f"⌀{int(size['d'])}"
            elif "w" in size:
                size_label = f"{int(size['w'])}x{int(size['h'])}"
            else:
                size_label = "-"
                
            angle = elem.get("angle")
            name = f"Отвод {angle}°" if (kind == "bend" and angle) else ("Тройник" if kind == "tee" else "Переход")
            
            key = (kind, size_label, elem.get("sortamentRef", "unknown"))
            if key in fittings_data:
                fittings_data[key]["quantity"] += 1
            else:
                fittings_data[key] = {
                    "name": name,
                    "size": size_label,
                    "quantity": 1
                }
        elif el_type == "terminal":
            kind = elem.get("kind", "grille")
            model = elem.get("model", "Terminal")
            name = "Решетка приточно-вытяжная" if kind == "grille" else "Диффузор круглый"
            
            key = (kind, model)
            if key in terminals_data:
                terminals_data[key]["quantity"] += 1
            else:
                terminals_data[key] = {
                    "name": name,
                    "model": model,
                    "quantity": 1
                }
        elif el_type == "equipment":
            model = elem.get("model", "Equipment")
            size = elem.get("size", {"l": 1200, "w": 600, "h": 600})
            size_label = f"{int(size['l'])}x{int(size['w'])}x{int(size['h'])}"
            
            key = (model, size_label)
            if key in equipment_data:
                equipment_data[key]["quantity"] += 1
            else:
                equipment_data[key] = {
                    "name": "Приточно-вытяжная установка",
                    "model": model,
                    "size": size_label,
                    "quantity": 1
                }
        elif el_type == "radiator":
            model = elem.get("model", "Radiator")
            size_label = "100x600x800 мм"
            
            key = (model, size_label)
            if key in equipment_data:
                equipment_data[key]["quantity"] += 1
            else:
                equipment_data[key] = {
                    "name": "Радиатор отопления секционный",
                    "model": model,
                    "size": size_label,
                    "quantity": 1
                }
        elif el_type == "ac_ceiling":
            model = elem.get("model", "Radiator")
            size_label = "800x800x50 мм"
            
            key = (model, size_label)
            if key in equipment_data:
                equipment_data[key]["quantity"] += 1
            else:
                equipment_data[key] = {
                    "name": "Кондиционер потолочный кассетный",
                    "model": model,
                    "size": size_label,
                    "quantity": 1
                }
                
    # 2. Создание Excel через openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость спецификации"
    
    # Включаем сетку
    ws.views.sheetView[0].showGridLines = True
    
    # Стили
    font_title = Font(name="Segoe UI", size=14, bold=True, color="1A202C")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="2D3748")
    font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=9)
    
    fill_header = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    fill_sub_header = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_side = Side(border_style="thin", color="CBD5E0")
    border_cell = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Заголовок проекта
    ws["B2"] = "ВЕДОМОСТЬ МАТЕРИАЛОВ И ОБОРУДОВАНИЯ ВЕНТИЛЯЦИИ"
    ws["B2"].font = font_title
    ws["B3"] = f"Проект: {db_project.name} | Версия: {db_project.version}"
    ws["B3"].font = Font(name="Segoe UI", size=9, italic=True)
    
    current_row = 5
    
    # 2.1 Таблица воздуховодов
    ws.cell(row=current_row, column=2, value="1. Воздуховоды вентиляционные").font = font_section
    current_row += 1
    
    headers_ducts = ["№", "Наименование", "Размер (мм)", "Длина (м)", "Площадь (м²)", "Масса (кг)"]
    for col_idx, h in enumerate(headers_ducts, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
    
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    item_num = 1
    total_len = 0.0
    total_area = 0.0
    total_weight = 0.0
    
    for key, data in ducts_data.items():
        row_vals = [
            item_num,
            data["name"],
            data["size"],
            round(data["length"], 2),
            round(data["area"], 2),
            round(data["weight"], 2)
        ]
        
        total_len += data["length"]
        total_area += data["area"]
        total_weight += data["weight"]
        
        for col_idx, val in enumerate(row_vals, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            cell.alignment = align_center if col_idx in [2, 4] else (align_left if col_idx == 3 else align_right)
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        item_num += 1
        
    # Итоги воздуховодов
    if ducts_data:
        ws.cell(row=current_row, column=2, value="Итого").font = Font(name="Segoe UI", size=9, bold=True)
        ws.cell(row=current_row, column=2).alignment = align_center
        ws.cell(row=current_row, column=2).border = border_cell
        
        ws.cell(row=current_row, column=3, value="-").border = border_cell
        ws.cell(row=current_row, column=3).alignment = align_center
        ws.cell(row=current_row, column=4, value="-").border = border_cell
        ws.cell(row=current_row, column=4).alignment = align_center
        
        cell_len = ws.cell(row=current_row, column=5, value=round(total_len, 2))
        cell_len.font = Font(name="Segoe UI", size=9, bold=True)
        cell_len.border = border_cell
        cell_len.alignment = align_right
        
        cell_area = ws.cell(row=current_row, column=6, value=round(total_area, 2))
        cell_area.font = Font(name="Segoe UI", size=9, bold=True)
        cell_area.border = border_cell
        cell_area.alignment = align_right
        
        cell_weight = ws.cell(row=current_row, column=7, value=round(total_weight, 2))
        cell_weight.font = Font(name="Segoe UI", size=9, bold=True)
        cell_weight.border = border_cell
        cell_weight.alignment = align_right
        
        current_row += 2
    else:
        current_row += 1
        
    # 2.2 Таблица фасонных деталей
    ws.cell(row=current_row, column=2, value="2. Фасонные соединительные детали").font = font_section
    current_row += 1
    
    headers_fittings = ["№", "Наименование детали", "Размер стыка (мм)", "Количество (шт)", "Ед. изм."]
    for col_idx, h in enumerate(headers_fittings, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_sub_header
        cell.alignment = align_center
        cell.border = border_cell
        
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    fit_num = 1
    for key, data in fittings_data.items():
        row_vals = [fit_num, data["name"], data["size"], data["quantity"], "шт"]
        for col_idx, val in enumerate(row_vals, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            cell.alignment = align_center if col_idx in [2, 5, 6] else (align_left if col_idx == 3 else align_right)
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        fit_num += 1
        
    current_row += 2
    
    # 2.3 Таблица оборудования и решеток
    ws.cell(row=current_row, column=2, value="3. Вентиляционное оборудование и оконечные приборы").font = font_section
    current_row += 1
    
    headers_eq = ["№", "Наименование прибора/модели", "Характеристика/Размер", "Количество (шт)", "Ед. изм."]
    for col_idx, h in enumerate(headers_eq, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    eq_num = 1
    # Сначала оборудование
    for key, data in equipment_data.items():
        row_vals = [eq_num, data["name"] + " (" + data["model"] + ")", data["size"], data["quantity"], "шт"]
        for col_idx, val in enumerate(row_vals, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            cell.alignment = align_center if col_idx in [2, 6] else (align_left if col_idx in [3, 4] else align_right)
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        eq_num += 1
        
    # Затем решетки
    for key, data in terminals_data.items():
        row_vals = [eq_num, data["name"], data["model"], data["quantity"], "шт"]
        for col_idx, val in enumerate(row_vals, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            cell.alignment = align_center if col_idx in [2, 6] else (align_left if col_idx in [3, 4] else align_right)
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        eq_num += 1
        
    # Автоподгонка ширины колонок
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Пропускаем титульный столбец A
        if col_letter == "A":
            ws.column_dimensions[col_letter].width = 3
            continue
            
        for cell in col:
            # Не учитываем длинный объединенный титульный заголовок в B2
            if cell.row in [2, 3]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Отдаем файл как стрим
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"spec_{project_id}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

if __name__ == "__main__":
    uvicorn.run("server.main:app", host="127.0.0.1", port=8000, reload=True)
